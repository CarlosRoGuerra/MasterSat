"""
Relatório de simulação de fechamento (formato do sistema antigo).

Extraído de billing_closure.py (BE-03): esta metade do arquivo original não
tocava em banco — recebe só o ``dict`` de simulação já pronto (produzido por
``simulate_closure``) e o transforma em texto monoespaçado, XLSX ou PDF. Same
comportamento de antes, só em módulo próprio por não compartilhar
responsabilidade nenhuma com o cálculo/persistência do fechamento.

Texto monoespaçado, agrupado por INTERVENIENTE (quem paga o boleto), com um
bloco por veículo e uma linha por rastreador — um veículo pode ter mais de um
equipamento, e cada um tem a própria mensalidade.
"""
from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

_LARGURA = 96          # colunas do relatório
_SEP = '=' * _LARGURA
_EMPRESA = 'MASTERSAT COMERCIO E SERVIÇOS DE RASTREAMENTO LTDA'


def _v(valor: float) -> str:
    """Valor no padrão brasileiro, sem símbolo (o relatório é monoespaçado)."""
    return f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _d(valor: date | None) -> str:
    return valor.strftime('%d/%m/%Y') if valor else '//'


def _par(esquerda: str, direita: str, largura: int = _LARGURA) -> str:
    """Duas colunas na mesma linha: uma à esquerda, outra à direita."""
    espaco = max(1, largura - len(esquerda) - len(direita))
    return f'{esquerda}{" " * espaco}{direita}'


def _total(rotulo: str, valor: float, recuo: int = 40) -> str:
    """
    Linha de total: rótulo recuado e valor à direita.

    Os 12 caracteres finais alinham na MESMA coluna do valor das linhas de
    rastreador (8 + 74 + 12 = 94), senão a coluna de valores fica em degrau.
    """
    corpo = f'{rotulo:<42}{_v(valor):>12}'
    return ' ' * recuo + corpo


def _chave_veiculo(item: dict) -> object:
    """Sem veículo, cada contrato é seu próprio grupo."""
    return item.get('vehicle_id') or f'c{item["contract_id"]}'


def _plural(n: int, singular: str, plural: str) -> str:
    return f'{n} {singular if n == 1 else plural}'


def _no_mes(valor: date | None, mes_ref: str) -> bool:
    """A data cai no mês de referência ('MM/AAAA')?"""
    if not valor or '/' not in (mes_ref or ''):
        return False
    mes, _, ano = mes_ref.partition('/')
    return valor.month == int(mes) and valor.year == int(ano)


def _contagens(itens: list[dict], desinstalacoes: list[dict], mes_ref: str) -> dict[str, int]:
    """Veículos, equipamentos, instalações e desinstalações de um conjunto."""
    return {
        'veiculos': len({_chave_veiculo(i) for i in itens}),
        # Um veículo pode ter mais de um rastreador — por isso a contagem é
        # separada da de veículos (ex.: ACQUE, 1 veículo com 2 equipamentos).
        'equipamentos': len({i['tracker_imei'] for i in itens if i.get('tracker_imei')}),
        'instalacoes': sum(1 for i in itens if _no_mes(i.get('tracker_install_date'), mes_ref)),
        'desinstalacoes': len(desinstalacoes),
    }


_MESES_EXT = ['', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
              'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']


def _mes_extenso(mes_ref: str) -> str:
    """'08/2026' → 'AGOSTO/2026'; devolve o original se não bater o formato."""
    mm, _, aaaa = (mes_ref or '').partition('/')
    if aaaa and mm.isdigit() and 1 <= int(mm) <= 12:
        return f'{_MESES_EXT[int(mm)]}/{aaaa}'
    return mes_ref or ''


def _totais_financeiros(itens: list[dict], simulation: dict) -> dict[str, float]:
    """Somatórios do fechamento, por natureza — a mesma conta no topo e no rodapé."""
    mensalidades = sum(float(i.get('billing_amount') or 0) for i in itens)
    produtos = sum(
        float(p.get('amount') or 0)
        for i in itens for p in (i.get('first_month_charges') or [])
    )
    taxas = float(simulation.get('total_uninstall_fees') or 0)
    servicos = float(simulation.get('total_services') or 0)
    return {
        'mensalidades': mensalidades, 'produtos': produtos,
        'taxas': taxas, 'servicos': servicos,
        'geral': mensalidades + produtos + taxas + servicos,
    }


def _painel_totais(contagem: dict[str, int], fin: dict[str, float]) -> list[str]:
    """
    Painel de totais no topo do relatório — a quantidade de tudo antes do
    detalhamento por cliente (pedido de 08/08/2026). Tabela de 5 colunas em
    caixa, no monoespaçado do relatório.
    """
    colunas = [
        ('Veículos', str(contagem['veiculos'])),
        ('Rastreadores', str(contagem['equipamentos'])),
        ('Instalações', str(contagem['instalacoes'])),
        ('Desinstalações', str(contagem['desinstalacoes'])),
        ('Total Geral', f'R$ {_v(fin["geral"])}'),
    ]
    w = 18  # 5 células de 18 + 6 bordas = 96 = _LARGURA
    borda = '+' + '+'.join('-' * w for _ in colunas) + '+'

    def _linha(celulas: list[str]) -> str:
        return '|' + '|'.join(c.center(w) for c in celulas) + '|'

    return [
        borda,
        _linha([rot for rot, _ in colunas]),
        borda,
        _linha([val for _, val in colunas]),
        borda,
    ]


def montar_linhas_simulacao(simulation: dict) -> list[str]:
    """
    Monta o relatório como lista de linhas de texto.

    Separado da geração do PDF para poder ser testado sem abrir o arquivo.
    """
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    mes_ref = simulation.get('reference_month', '')
    titulo = f'PRÉVIA DE FECHAMENTO — {_mes_extenso(mes_ref)}'.strip(' —')
    linhas: list[str] = [_SEP, titulo.center(_LARGURA), _SEP]

    itens = simulation.get('items') or []
    if not itens:
        linhas += ['', 'Nenhum contrato a faturar no período.']
        return linhas

    # 1) agrupa por interveniente  2) dentro dele, por veículo
    por_interveniente: dict[str, list[dict]] = defaultdict(list)
    for item in itens:
        por_interveniente[item.get('interveniente_nome') or item['client_name']].append(item)

    # As desinstalações vêm por cliente; o relatório agrupa por interveniente.
    # Este mapa leva uma à outra para a contagem cair no bloco certo.
    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }
    desinst_por_grupo: dict[str, list[dict]] = defaultdict(list)
    for ev in simulation.get('uninstall_events') or []:
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        desinst_por_grupo[grupo].append(ev)

    fin = _totais_financeiros(itens, simulation)

    # ── Painel de totais no topo (a quantidade de tudo antes dos clientes) ──
    linhas.append(_par(f'Mês de referência: {mes_ref}', f'Gerado em {agora}'))
    linhas.append('')
    linhas += _painel_totais(
        _contagens(itens, simulation.get('uninstall_events') or [], mes_ref), fin)
    linhas.append('')

    for interveniente, do_grupo in sorted(por_interveniente.items()):
        por_veiculo: dict[object, list[dict]] = defaultdict(list)
        for item in do_grupo:
            por_veiculo[_chave_veiculo(item)].append(item)

        venc = do_grupo[0].get('due_date')
        mes_venc = venc.strftime('%m/%Y') if venc else ''
        qtd = _contagens(do_grupo, desinst_por_grupo.get(interveniente, []), mes_ref)

        linhas.append(f'INTERVENIENTE: {interveniente}')
        linhas.append(f'MATRIZ/FILIAL: {_EMPRESA}')
        linhas.append('')
        linhas.append(_par(
            f'MÊS REFERENTE: {mes_ref}',
            _par(f'MÊS VENCIMENTO: {mes_venc}',
                 f'QUANTIDADE VEÍCULOS: {qtd["veiculos"]}', 62),
        ))
        linhas.append(_par(
            f'QUANTIDADE EQUIPAMENTOS: {qtd["equipamentos"]}',
            _par(f'INSTALAÇÕES NO MÊS: {qtd["instalacoes"]}',
                 f'DESINSTALAÇÕES NO MÊS: {qtd["desinstalacoes"]}', 62),
        ))
        linhas.append('')

        total_grupo = 0.0
        for contratos in por_veiculo.values():
            primeiro = contratos[0]
            venc_v = primeiro.get('due_date')
            dia = primeiro.get('billing_day') or (venc_v.day if venc_v else '')

            linhas.append(f'CLIENTE: {primeiro["client_name"]}')
            linhas.append(_par(
                f'PLACA: {primeiro.get("vehicle_plate") or "—"}',
                f'TIPO VEÍCULO: {(primeiro.get("vehicle_type") or "—").upper()}', 78))
            linhas.append(_par(
                f'DATA CADASTRO: {_d(primeiro.get("vehicle_created_at"))}',
                f'RASTREADOR: {primeiro.get("tracker_imei") or "—"}', 78))
            linhas.append(_par(
                f'DATA CONTRATO: {_d(primeiro.get("contract_start_date"))}',
                f'VENCIMENTO: {dia}[{_d(venc_v)}]', 78))
            linhas.append('')

            total_veiculo = 0.0
            for c in contratos:
                mensalidade = float(c.get('billing_amount') or 0)
                instalacao = _d(c.get('tracker_install_date'))
                rotulo = f'RASTREADOR: DATA INSTALAÇÃO [{instalacao}] - MENSALIDADE {_v(mensalidade)}:'
                linhas.append(' ' * 8 + f'{rotulo:<74}{_v(mensalidade):>12}'.rstrip())

                # Serviços/produtos embutidos na primeira cobrança
                produtos = c.get('first_month_charges') or []
                soma_produtos = 0.0
                for prod in produtos:
                    val = float(prod.get('amount') or 0)
                    soma_produtos += val
                    desc = f'PRODUTO - {str(prod.get("title") or "").upper()}:'
                    linhas.append(' ' * 8 + f'{desc:<74}{_v(val):>12}'.rstrip())
                if produtos:
                    linhas.append(_total('SOMA PRODUTOS:', soma_produtos))

                linhas.append(_total('TOTAL RASTREADOR:', mensalidade))
                linhas.append('')
                total_veiculo += mensalidade + soma_produtos

            linhas.append(_total('TOTAL VEÍCULO:', total_veiculo))
            linhas.append('')
            total_grupo += total_veiculo

        # Sem impostos configurados, os dois totais são iguais — as duas linhas
        # existem porque o relatório de referência as traz.
        linhas.append(_total('TOTAL BOLETO S/ IMPOSTOS:', total_grupo))
        linhas.append(_total('TOTAL BOLETO C/ IMPOSTOS:', total_grupo))
        linhas.append(_SEP)

    linhas += _movimentacao_do_mes(itens, simulation, por_interveniente, mes_ref)
    linhas += _resumo_geral(simulation, itens, por_interveniente, mes_ref, fin)
    return linhas


def _movimentacao_do_mes(itens: list[dict], simulation: dict,
                         por_interveniente: dict[str, list[dict]],
                         mes_ref: str) -> list[str]:
    """
    Instalações e desinstalações do período, uma a uma.

    Pedido da reunião de 07/08/2026: o resumo diz quantas foram; aqui se vê
    quais — por interveniente e por cliente, com data, placa e equipamento.
    """
    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }

    instalacoes: dict[str, list[dict]] = defaultdict(list)
    for item in itens:
        if _no_mes(item.get('tracker_install_date'), mes_ref):
            instalacoes[item.get('interveniente_nome') or item['client_name']].append(item)

    desinstalacoes: dict[str, list[dict]] = defaultdict(list)
    for ev in simulation.get('uninstall_events') or []:
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        desinstalacoes[grupo].append(ev)

    if not instalacoes and not desinstalacoes:
        return []

    linhas = ['', f'MOVIMENTAÇÃO DO PERÍODO — {mes_ref}'.center(_LARGURA), _SEP]

    for grupo in sorted(set(instalacoes) | set(desinstalacoes)):
        linhas.append(f'INTERVENIENTE: {grupo}')

        for item in sorted(instalacoes.get(grupo, []),
                           key=lambda i: (i.get('tracker_install_date') or date.min,
                                          i.get('vehicle_plate') or '')):
            linhas.append(_par(
                f'    INSTALAÇÃO  {_d(item.get("tracker_install_date"))}  '
                f'{(item.get("vehicle_plate") or "SEM PLACA"):<10} '
                f'RASTREADOR {item.get("tracker_imei") or "—"}',
                item['client_name'][:34],
            ))

        for ev in sorted(desinstalacoes.get(grupo, []),
                         key=lambda e: (e.get('uninstall_date') or date.min,
                                        e.get('vehicle_plate') or '')):
            taxa = float(ev.get('fee_amount') or 0)
            # O valor pequeno continua devido e será acumulado; o relatório não
            # pode apresentá-lo como isento ou descartado.
            sufixo = (
                f'TAXA {_v(taxa)} (AGUARDANDO ACUMULAÇÃO)'
                if ev.get('deferred') else f'TAXA {_v(taxa)}'
            )
            linhas.append(_par(
                f'    DESINSTALAÇÃO  {_d(ev.get("uninstall_date"))}  '
                f'{(ev.get("vehicle_plate") or "SEM PLACA"):<10} {sufixo}',
                (ev.get('client_name') or '')[:34],
            ))

        linhas.append('    Subtotal: ' + ' · '.join((
            _plural(len(instalacoes.get(grupo, [])), 'instalação', 'instalações'),
            _plural(len(desinstalacoes.get(grupo, [])), 'desinstalação', 'desinstalações'),
        )))
        linhas.append('')

    linhas.append(_SEP)
    return linhas


def _resumo_geral(simulation: dict, itens: list[dict],
                  por_interveniente: dict[str, list[dict]], mes_ref: str,
                  fin: dict[str, float]) -> list[str]:
    """Fecha o relatório com o consolidado do período (mesma conta do topo)."""
    desinstalacoes = simulation.get('uninstall_events') or []
    qtd = _contagens(itens, desinstalacoes, mes_ref)

    linhas = ['', f'RESUMO DO FECHAMENTO — {mes_ref}'.center(_LARGURA), _SEP]
    linhas.append(_par(
        f'INTERVENIENTES: {len(por_interveniente)}',
        _par(f'VEÍCULOS: {qtd["veiculos"]}',
             f'EQUIPAMENTOS: {qtd["equipamentos"]}', 62),
    ))
    linhas.append(_par(
        f'INSTALAÇÕES NO MÊS: {qtd["instalacoes"]}',
        _par(f'DESINSTALAÇÕES NO MÊS: {qtd["desinstalacoes"]}',
             f'CONTRATOS: {len(itens)}', 62),
    ))

    # Cobranças já geradas continuam listadas no detalhamento; sem a linha, o
    # total do resumo pareceria não bater com o que o fechamento vai gerar.
    ja_geradas = int(simulation.get('already_generated') or 0)
    if ja_geradas:
        linhas.append(f'CONTRATOS JÁ FATURADOS NO PERÍODO (inclusos acima): {ja_geradas}')

    linhas.append('')
    linhas.append(_total('TOTAL MENSALIDADES:', fin['mensalidades']))
    if fin['produtos']:
        linhas.append(_total('TOTAL PRODUTOS/SERVIÇOS NA 1ª COBRANÇA:', fin['produtos']))
    if fin['taxas']:
        linhas.append(_total('TOTAL TAXAS DE DESINSTALAÇÃO:', fin['taxas']))
    if fin['servicos']:
        linhas.append(_total('TOTAL SERVIÇOS AVULSOS:', fin['servicos']))
    linhas.append(_total('TOTAL GERAL:', fin['geral']))
    linhas.append(_SEP)
    return linhas


def generate_closure_xlsx(simulation: dict) -> BytesIO:
    """
    Simulação de fechamento em Excel (.xlsx), com 3 abas:
      Resumo        — período, painel de totais e os somatórios financeiros
      Contratos     — uma linha por contrato (interveniente, cliente, veículo…)
      Movimentação  — instalações e desinstalações do período, uma por linha

    Ao contrário do PDF (texto monoespaçado), aqui os dados vão em colunas para
    a operação filtrar, ordenar e somar na planilha.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    mes_ref = simulation.get('reference_month', '')
    itens = simulation.get('items') or []
    desinst = simulation.get('uninstall_events') or []
    fin = _totais_financeiros(itens, simulation)
    cont = _contagens(itens, desinst, mes_ref)
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }

    MONEY = '"R$" #,##0.00'
    DATA = 'DD/MM/YYYY'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2A44')
    bold = Font(bold=True)

    def _cabecalho(ws, colunas: list[str]) -> None:
        for ci, nome in enumerate(colunas, 1):
            cell = ws.cell(1, ci, nome)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'

    wb = Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumo'
    ws['A1'] = 'PRÉVIA DE FECHAMENTO'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Mês de referência: {mes_ref}'
    ws['A3'] = f'Gerado em: {agora}'

    ws['A5'] = 'Totais do período'
    ws['A5'].font = bold
    linha = 6
    for rotulo, valor in (
        ('Veículos', cont['veiculos']),
        ('Rastreadores', cont['equipamentos']),
        ('Instalações', cont['instalacoes']),
        ('Desinstalações', cont['desinstalacoes']),
    ):
        ws.cell(linha, 1, rotulo).font = bold
        ws.cell(linha, 2, valor)
        linha += 1

    linha += 1
    ws.cell(linha, 1, 'Totais financeiros').font = bold
    linha += 1
    for rotulo, valor in (
        ('Total mensalidades', fin['mensalidades']),
        ('Total produtos/serviços (1ª cobrança)', fin['produtos']),
        ('Total taxas de desinstalação', fin['taxas']),
        ('Total serviços avulsos', fin['servicos']),
        ('TOTAL GERAL', fin['geral']),
    ):
        c1 = ws.cell(linha, 1, rotulo)
        c2 = ws.cell(linha, 2, float(valor))
        c2.number_format = MONEY
        if rotulo.startswith('TOTAL GERAL'):
            c1.font = bold
            c2.font = bold
        linha += 1
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    # ── Aba Contratos ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Contratos')
    colunas = [
        'Interveniente', 'Cliente', 'Placa', 'Tipo veículo', 'Rastreador (IMEI)',
        'Data instalação', 'Data contrato', 'Plano', 'Mensalidade',
        'Produtos 1ª cobrança', 'Total', 'Vencimento', 'Período', 'Já faturado',
    ]
    _cabecalho(ws2, colunas)
    rr = 2
    for it in itens:
        produtos = sum(float(p.get('amount') or 0) for p in (it.get('first_month_charges') or []))
        valores = [
            it.get('interveniente_nome') or it.get('client_name'),
            it.get('client_name'),
            it.get('vehicle_plate') or '',
            (it.get('vehicle_type') or '').upper(),
            it.get('tracker_imei') or '',
            it.get('tracker_install_date'),
            it.get('contract_start_date'),
            it.get('plan_name'),
            float(it.get('billing_amount') or 0),
            produtos,
            float(it.get('total_first_billing') or 0),
            it.get('due_date'),
            it.get('period_label') or '',
            'Sim' if it.get('already_generated') else 'Não',
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws2.cell(rr, ci, val)
            if ci in (6, 7, 12):
                cell.number_format = DATA
            elif ci in (9, 10, 11):
                cell.number_format = MONEY
        rr += 1
    for ci, largura in enumerate([28, 28, 10, 14, 20, 15, 15, 20, 14, 18, 14, 13, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = largura

    # ── Aba Movimentação ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Movimentação')
    _cabecalho(ws3, ['Tipo', 'Data', 'Interveniente', 'Cliente', 'Placa', 'Rastreador', 'Taxa', 'Status'])
    rr = 2
    for it in itens:
        if _no_mes(it.get('tracker_install_date'), mes_ref):
            valores = [
                'Instalação', it.get('tracker_install_date'),
                it.get('interveniente_nome') or it.get('client_name'),
                it.get('client_name'), it.get('vehicle_plate') or '',
                it.get('tracker_imei') or '', None, '',
            ]
            for ci, val in enumerate(valores, 1):
                cell = ws3.cell(rr, ci, val)
                if ci == 2:
                    cell.number_format = DATA
            rr += 1
    for ev in desinst:
        taxa = float(ev.get('fee_amount') or 0)
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        valores = [
            'Desinstalação', ev.get('uninstall_date'), grupo,
            ev.get('client_name') or '', ev.get('vehicle_plate') or '', '', taxa,
            ('Aguardando acumulação' if ev.get('deferred') else 'A faturar'),
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws3.cell(rr, ci, val)
            if ci == 2:
                cell.number_format = DATA
            elif ci == 7 and val is not None:
                cell.number_format = MONEY
        rr += 1
    for ci, largura in enumerate([14, 12, 28, 28, 10, 20, 12, 24], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = largura

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_closure_pdf(simulation: dict) -> BytesIO:
    """Simulação de fechamento em PDF, no formato monoespaçado do SGR."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas

    linhas = montar_linhas_simulacao(simulation)

    buffer = BytesIO()
    c = pdfcanvas.Canvas(buffer, pagesize=A4)
    c.setTitle('Simulação de Fechamento — MasterSat')
    largura, altura = A4
    margem = 10 * mm
    topo = altura - margem
    passo = 3.3 * mm
    fonte = 6.5

    y = topo
    pagina = 1
    for linha in linhas:
        if y < margem + 8 * mm:
            c.setFont('Courier', 6)
            c.drawCentredString(largura / 2, margem, f'Página {pagina}')
            c.showPage()
            pagina += 1
            y = topo
        c.setFont('Courier', fonte)
        c.drawString(margem, y, linha)
        y -= passo

    c.setFont('Courier', 6)
    c.drawCentredString(largura / 2, margem, f'Página {pagina}')
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer
