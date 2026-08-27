from __future__ import annotations

from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO

from sqlalchemy import func, or_, select, text
from sqlalchemy.orm import Session, aliased

from app.core.timezone import hoje
from app.models.billing import Billing
from app.models.client import Client
from app.models.client_charge_item import ClientChargeItem
from app.models.contract import Contract
from app.models.enums import BillingStatus
from app.models.plan import Plan
from app.models.service_product import ServiceProduct
from app.models.tracker import Tracker
from app.models.uninstall_event import UninstallEvent
from app.models.vehicle import Vehicle
from app.services.financial import (
    _quantize_amount,
    add_months,
    associate_billing_charge_item,
    charge_item_effective_billing_count,
    contract_payer_client_id,
    decimal_to_float,
    generate_item_billings,
    normalize_due_date,
    plan_title,
    period_label_for_date,
    refresh_overdue_statuses,
)

MIN_BILLING_AMOUNT = Decimal('5.00')


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _billing_due_in_month(contract: Contract, plan: Plan, reference_month: date) -> date | None:
    interval = max(int(getattr(plan, 'billing_interval_months', 1) or 1), 1)
    billing_day = contract.billing_day or 1
    months_since_start = (
        (reference_month.year - contract.start_date.year) * 12
        + (reference_month.month - contract.start_date.month)
    )
    if months_since_start < 0:
        return None
    cycle = months_since_start // interval
    due_date = normalize_due_date(contract.start_date, cycle, billing_day, interval)
    # Never produce a due date that predates the contract start (billing_day < start day)
    if due_date < contract.start_date:
        return None
    if due_date.year == reference_month.year and due_date.month == reference_month.month:
        return due_date
    return None


def _lock_competencia(db: Session, reference_month: date) -> None:
    """Serializa fechamentos concorrentes da MESMA competência.

    Sem isto, dois fechamentos simultâneos do mesmo mês veem ``already_generated``
    False ao mesmo tempo e ambos inserem — duplicando as cobranças. O lock de
    transação (Postgres) segura o segundo até o primeiro comitar; em SQLite
    (testes, serial) é no-op.
    """
    if db.bind is None or db.bind.dialect.name != 'postgresql':
        return
    # Chave estável por competência (AAAAMM) — meses diferentes não se bloqueiam.
    chave = reference_month.year * 100 + reference_month.month
    db.execute(text('SELECT pg_advisory_xact_lock(:k)'), {'k': chave})


def _has_existing_billing(db: Session, contract_id: int, period_label: str) -> bool:
    return db.query(Billing).filter(
        Billing.is_deleted.is_(False),
        Billing.contract_id == contract_id,
        Billing.period_label == period_label,
        # 'carne': parcela de carnê já cobre o mês — sem isto o fechamento
        # mensal gerava uma mensalidade recorrente POR CIMA de um mês já
        # pago via carnê (cobrança duplicada).
        Billing.billing_type.in_(['recorrente', 'prorata', 'primeira_mensalidade', 'carne']),
    ).first() is not None


def _prorata_fields(plan_price: float, start_date: date) -> tuple[bool, float, int, int]:
    """
    Retorna (is_prorata, billing_amount, prorated_days, days_in_month).
    Pro-rata se aplica apenas quando o contrato começa após o dia 1 do mês.
    """
    if start_date.day == 1:
        return False, plan_price, 0, 0
    dim = monthrange(start_date.year, start_date.month)[1]
    remaining = dim - start_date.day + 1
    prorated = float(
        (Decimal(str(plan_price)) * Decimal(remaining) / Decimal(dim)).quantize(
            Decimal('0.01'), rounding=ROUND_HALF_UP
        )
    )
    return True, prorated, remaining, dim


def _apply_client_scope(query, client_column, filter_type: str, client_id: int | None):
    """Aplica o mesmo recorte de clientes a qualquer categoria do fechamento."""
    eligible_clients = select(Client.id).where(Client.is_deleted.is_(False))
    if filter_type in ('pf', 'pj'):
        eligible_clients = eligible_clients.where(Client.type == filter_type)
    elif filter_type == 'client' and client_id is not None:
        eligible_clients = eligible_clients.where(Client.id == client_id)
    return query.filter(client_column.in_(eligible_clients))


def _pending_uninstall_events_for_month(
    db: Session,
    reference_month: date,
    filter_type: str = 'all',
    client_id: int | None = None,
) -> list[UninstallEvent]:
    """Eventos vencidos até a competência, inclusive os esquecidos em meses anteriores.

    ``skipped`` era o estado terminal usado para valores abaixo de R$ 5. Ele é
    incluído para recuperar dados históricos e volta a ``pending`` enquanto
    aguarda acumulação suficiente.
    """
    if reference_month.month == 12:
        month_end = date(reference_month.year + 1, 1, 1)
    else:
        month_end = date(reference_month.year, reference_month.month + 1, 1)
    query = (
        db.query(UninstallEvent)
        .filter(
            UninstallEvent.status.in_(('pending', 'skipped')),
            UninstallEvent.billing_id.is_(None),
            UninstallEvent.uninstall_date < month_end,
        )
    )
    return _apply_client_scope(
        query, UninstallEvent.client_id, filter_type, client_id,
    ).order_by(UninstallEvent.uninstall_date.asc(), UninstallEvent.id.asc()).all()


def uninstall_fee_for_event(db: Session, event: UninstallEvent) -> tuple[Decimal, str]:
    """Valor e título da taxa de um evento de desinstalação.

    ``fee_amount`` é o valor efetivamente acordado no momento da retirada e
    tem precedência absoluta: o produto de serviço define apenas O QUE foi
    cobrado (título/vínculo com o catálogo), nunca QUANTO.

    Antes as duas coisas eram somadas. Como a tela preenche a taxa direta com
    o preço do produto ao selecioná-lo, escolher um serviço de desinstalação
    cobrava o dobro de forma determinística. Somar também deixava o valor
    refém do catálogo: mudar o preço do produto depois alterava uma cobrança
    já negociada com o cliente.

    O preço do produto só é consultado como fallback, para eventos antigos
    gravados sem ``fee_amount``.
    """
    product = None
    if event.service_product_id:
        candidate = db.get(ServiceProduct, event.service_product_id)
        if candidate and not candidate.is_deleted:
            product = candidate

    if event.fee_amount is not None and Decimal(str(event.fee_amount)) > 0:
        fee_amount = Decimal(str(event.fee_amount))
    elif product is not None:
        fee_amount = Decimal(str(product.default_price))
    else:
        fee_amount = Decimal('0')

    return fee_amount, (product.name if product else 'Taxa de desinstalação')


def _due_date_for_uninstall_event(event: UninstallEvent, db: Session) -> date:
    contract = db.get(Contract, event.contract_id) if event.contract_id else None
    client = db.get(Client, event.client_id)
    billing_day = (
        (contract.billing_day if contract and contract.billing_day else None)
        or (client.billing_day if client and client.billing_day else None)
        or 1
    )
    dim = monthrange(event.uninstall_date.year, event.uninstall_date.month)[1]
    fee_billing_day = min(billing_day, dim)
    due = date(event.uninstall_date.year, event.uninstall_date.month, fee_billing_day)
    if due <= event.uninstall_date:
        due = add_months(due, 1)
    return due


def _uninstall_event_payer_client_id(db: Session, event: UninstallEvent) -> int:
    """Pagador do evento, sem aceitar referência contratual inconsistente."""
    if event.payer_client_id:
        payer = db.get(Client, event.payer_client_id)
        if not payer or payer.is_deleted:
            raise ValueError(
                f'Responsável financeiro #{event.payer_client_id} do evento '
                f'#{event.id} não está disponível.'
            )
        return payer.id
    if not event.contract_id:
        return event.client_id
    contract = db.get(Contract, event.contract_id)
    if not contract or contract.is_deleted:
        raise ValueError(f'Contrato #{event.contract_id} da desinstalação não está disponível.')
    if contract.client_id != event.client_id:
        raise ValueError(
            f'Evento #{event.id} e contrato #{contract.id} pertencem a clientes diferentes.'
        )
    return contract_payer_client_id(db, contract)


def _first_cycle_charge_items(
    db: Session,
    client_id: int,
    contract_id: int,
    reference_month: date,
) -> list[dict]:
    """
    Retorna ChargeItems de parcela única (installment_count=1) que:
    - Pertencem explicitamente ao contrato (itens genéricos ficam avulsos)
    - Têm start_date dentro do mês de referência (serviços do início do contrato)
    - Ainda não foram faturados (active=True e billing_count=0)

    Esses itens serão embutidos na cobrança combinada da primeira mensalidade,
    em vez de gerar faturas separadas.
    """
    month_start = reference_month.replace(day=1)
    if reference_month.month == 12:
        month_end = date(reference_month.year + 1, 1, 1)
    else:
        month_end = date(reference_month.year, reference_month.month + 1, 1)

    q = db.query(ClientChargeItem).filter(
        ClientChargeItem.is_deleted.is_(False),
        ClientChargeItem.active.is_(True),
        ClientChargeItem.client_id == client_id,
        ClientChargeItem.contract_id == contract_id,
        ClientChargeItem.installment_count == 1,
        ClientChargeItem.start_date >= month_start,
        ClientChargeItem.start_date < month_end,
    )
    result = []
    for item in q.all():
        billing_count = charge_item_effective_billing_count(db, item.id)
        if billing_count > 0:
            continue  # já faturado por outro caminho

        result.append({
            'item_id': item.id,
            'title': item.title,
            'amount': float(_quantize_amount(item.total_amount)),
        })

    return result


def _pending_charge_items(
    db: Session,
    reference_month: date,
    exclude_ids: set[int] | None = None,
    filter_type: str = 'all',
    client_id: int | None = None,
) -> list[dict]:
    """
    Retorna ClientChargeItems ativos cujos billings ainda não foram totalmente gerados
    e cujo start_date é anterior ao final do mês de referência.
    exclude_ids: item_ids já embutidos em cobranças combinadas de primeiro mês.
    """
    if reference_month.month == 12:
        month_end = date(reference_month.year + 1, 1, 1)
    else:
        month_end = date(reference_month.year, reference_month.month + 1, 1)

    query = (
        db.query(ClientChargeItem)
        .filter(
            ClientChargeItem.is_deleted.is_(False),
            ClientChargeItem.active.is_(True),
            ClientChargeItem.start_date < month_end,
        )
    )
    items = _apply_client_scope(
        query, ClientChargeItem.client_id, filter_type, client_id,
    ).order_by(ClientChargeItem.id.asc()).all()

    result = []
    for item in items:
        if exclude_ids and item.id in exclude_ids:
            continue

        billing_count = charge_item_effective_billing_count(db, item.id)

        installments = max(int(item.installment_count or 1), 1)
        if billing_count >= installments:
            continue

        client = db.get(Client, item.client_id)
        vehicle = db.get(Vehicle, item.vehicle_id) if item.vehicle_id else None
        remaining = installments - billing_count
        total = Decimal(str(item.total_amount))
        per_installment = (total / installments).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)

        result.append({
            'type': 'servico',
            'item_id': item.id,
            'client_id': item.client_id,
            'client_name': client.name if client else f'Cliente #{item.client_id}',
            'client_type': client.type if client else 'pf',
            'vehicle_plate': vehicle.plate if vehicle else None,
            'title': item.title,
            'installment_count': installments,
            'generated_count': billing_count,
            'remaining_installments': remaining,
            'per_installment_amount': float(per_installment),
            'total_remaining': float(per_installment * remaining),
            'start_date': item.start_date,
        })

    return result


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def simulate_closure(
    db: Session,
    reference_month: date,
    filter_type: str = 'all',
    client_id: int | None = None,
    *,
    commit: bool = True,
) -> dict:
    # commit=False quando chamada de dentro de execute_closure: o refresh abaixo
    # comitava e, com isso, encerrava a transação que segura o lock da
    # competência — na prática o lock morria aqui, antes de qualquer cobrança
    # ser criada.
    refresh_overdue_statuses(db, commit=commit)

    # Cliente que responde pela cobrança (interveniente). Sem ele, o próprio
    # cliente do contrato é o responsável — é por ele que o relatório agrupa.
    Interveniente = aliased(Client)

    query = db.query(Contract, Client, Plan, Vehicle, Tracker, Interveniente).join(
        Client, Client.id == Contract.client_id
    ).join(
        Plan, Plan.id == Contract.plan_id
    ).outerjoin(
        Vehicle, Vehicle.id == Contract.vehicle_id
    ).outerjoin(
        Tracker, Tracker.id == Contract.tracker_id
    ).outerjoin(
        Interveniente, Interveniente.id == Contract.interveniente_client_id
    ).filter(
        Contract.is_deleted.is_(False),
        Contract.status == 'ativo',
        Client.is_deleted.is_(False),
        Plan.is_deleted.is_(False),
        # Contrato com vigência já encerrada antes do mês de referência não entra
        # no fechamento — senão sairia boleto/NFS-e de um contrato que acabou.
        or_(Contract.end_date.is_(None), Contract.end_date >= reference_month),
    )

    if filter_type == 'pf':
        query = query.filter(Client.type == 'pf')
    elif filter_type == 'pj':
        query = query.filter(Client.type == 'pj')
    elif filter_type == 'client' and client_id:
        query = query.filter(Client.id == client_id)

    items = []
    for contract, client, plan, vehicle, tracker, interveniente in query.all():
        due_date = _billing_due_in_month(contract, plan, reference_month)
        if due_date is None:
            continue

        interval = max(int(getattr(plan, 'billing_interval_months', 1) or 1), 1)
        period_label = period_label_for_date(due_date, interval)
        already = _has_existing_billing(db, contract.id, period_label)
        plan_price = decimal_to_float(plan.price)

        # First billing month: normally the start month, but shifts to the NEXT month
        # when billing_day < start_date.day (that calendar day has already passed).
        _billing_day = contract.billing_day or 1
        if _billing_day >= contract.start_date.day:
            _first_billing_month = contract.start_date.replace(day=1)
        else:
            _first_billing_month = add_months(contract.start_date.replace(day=1), 1)

        first_cycle = (
            reference_month.year == _first_billing_month.year
            and reference_month.month == _first_billing_month.month
        )
        if first_cycle:
            is_prorata, billing_amount, prorated_days, days_in_month = _prorata_fields(
                plan_price, contract.start_date
            )
            # Charge items with start_date in the contract's start month are embedded
            first_charges = _first_cycle_charge_items(
                db, client.id, contract.id,
                contract.start_date.replace(day=1),
            )
        else:
            is_prorata, billing_amount, prorated_days, days_in_month = False, plan_price, 0, 0
            first_charges = []

        total_first_billing = billing_amount + sum(c['amount'] for c in first_charges)

        items.append({
            'type': 'recorrente',
            'contract_id': contract.id,
            'client_id': client.id,
            'client_name': client.name,
            'payer_client_id': contract_payer_client_id(db, contract),
            'payer_name': (
                interveniente.name if interveniente and not interveniente.is_deleted else client.name
            ),
            'client_type': client.type,
            'vehicle_plate': vehicle.plate if vehicle else None,
            'tracker_imei': tracker.imei if tracker else None,
            'plan_name': plan.name,
            'plan_price': plan_price,
            'billing_amount': billing_amount,
            'is_prorata': is_prorata,
            'prorated_days': prorated_days,
            'days_in_month': days_in_month,
            'first_month_charges': first_charges,
            'total_first_billing': total_first_billing,
            'period_label': period_label,
            'due_date': due_date,
            'already_generated': already,
            'billing_day': contract.billing_day,
            # Campos usados pelo relatório de simulação (formato do SGR):
            # agrupa por interveniente e detalha veículo + rastreadores.
            'interveniente_nome': (interveniente.name if interveniente else client.name),
            'vehicle_id': vehicle.id if vehicle else None,
            'vehicle_type': (vehicle.type if vehicle else None),
            'vehicle_created_at': (
                vehicle.created_at.date() if vehicle and getattr(vehicle, 'created_at', None) else None
            ),
            'contract_start_date': contract.start_date,
            'tracker_install_date': (tracker.install_date if tracker else None),
        })

    # Exclui dos serviços avulsos os itens já embutidos em cobranças de primeiro mês
    embedded_ids: set[int] = {
        c['item_id']
        for item in items
        for c in item['first_month_charges']
    }

    # Eventos de desinstalação vencidos até esta competência. Valores pequenos
    # são acumulados por cliente; nunca mais viram perda terminal em ``skipped``.
    uninstall_events = _pending_uninstall_events_for_month(
        db, reference_month, filter_type, client_id,
    )
    uninstall_amounts = {
        event.id: uninstall_fee_for_event(db, event)[0]
        for event in uninstall_events
    }
    payer_ids_by_event = {
        event.id: _uninstall_event_payer_client_id(db, event)
        for event in uninstall_events
    }
    totals_by_payer: dict[int, Decimal] = defaultdict(lambda: Decimal('0.00'))
    for event in uninstall_events:
        totals_by_payer[payer_ids_by_event[event.id]] += uninstall_amounts[event.id]

    uninstall_items = []
    for event in uninstall_events:
        client = db.get(Client, event.client_id)
        vehicle = db.get(Vehicle, event.vehicle_id)
        payer_id = payer_ids_by_event[event.id]
        payer = db.get(Client, payer_id)
        fee_amount = uninstall_amounts[event.id]
        aggregation_total = totals_by_payer[payer_id]
        deferred = aggregation_total < MIN_BILLING_AMOUNT
        uninstall_items.append({
            'type': 'taxa_desinstalacao',
            'event_id': event.id,
            'client_id': event.client_id,
            'client_name': client.name if client else f'Cliente #{event.client_id}',
            'payer_client_id': payer_id,
            'payer_name': payer.name if payer else f'Responsável financeiro #{payer_id}',
            'client_type': client.type if client else 'pf',
            'vehicle_plate': vehicle.plate if vehicle else None,
            'uninstall_date': event.uninstall_date,
            'fee_amount': float(fee_amount),
            'deferred': deferred,
            # Compatibilidade temporária para clientes antigos da API. Agora
            # significa "não faturado nesta rodada", sem mudar o evento para
            # um estado terminal.
            'skipped': deferred,
            'skip_reason': (
                f'Acumulado do cliente em R$ {float(aggregation_total):.2f}; '
                f'aguardando mínimo de R$ {float(MIN_BILLING_AMOUNT):.2f}'
                if deferred else None
            ),
            'aggregation_total': float(aggregation_total),
        })

    # Serviços / cobranças avulsas pendentes (exclui os embutidos)
    charge_items = _pending_charge_items(
        db, reference_month, exclude_ids=embedded_ids,
        filter_type=filter_type, client_id=client_id,
    )

    to_generate = [i for i in items if not i['already_generated']]
    already_done = [i for i in items if i['already_generated']]
    # total_amount inclui os serviços embutidos na primeira cobrança
    total_amount = sum(i['total_first_billing'] for i in to_generate)
    total_uninstall = sum(i['fee_amount'] for i in uninstall_items if not i['deferred'])
    total_services = sum(i['total_remaining'] for i in charge_items)

    return {
        'reference_month': reference_month.strftime('%m/%Y'),
        'total_contracts': len(items),
        'to_generate': len(to_generate),
        'already_generated': len(already_done),
        'total_amount': round(total_amount, 2),
        'items': items,
        'uninstall_events': uninstall_items,
        'total_uninstall_fees': round(total_uninstall, 2),
        'charge_items': charge_items,
        'total_services': round(total_services, 2),
        'grand_total': round(total_amount + total_uninstall + total_services, 2),
    }


def execute_closure(
    db: Session,
    reference_month: date,
    filter_type: str = 'all',
    client_id: int | None = None,
    contract_ids: list[int] | None = None,
    uninstall_event_ids: list[int] | None = None,
    charge_item_ids: list[int] | None = None,
) -> dict:
    # Trava a competência ANTES de simular: simulação + geração ficam atômicas em
    # relação a outro fechamento do mesmo mês, fechando a corrida de duplicação.
    # O lock é de TRANSAÇÃO (pg_advisory_xact_lock), então nada daqui até o
    # commit final pode comitar — daí o commit=False propagado abaixo.
    _lock_competencia(db, reference_month)
    simulation = simulate_closure(db, reference_month, filter_type, client_id, commit=False)
    # Ao receber qualquer lista de seleção, opera em modo snapshot/fail-closed:
    # categorias omitidas significam seleção vazia, não "processar tudo". Sem
    # lista alguma preservamos o comando administrativo de fechamento integral.
    exact_selection = any(
        ids is not None
        for ids in (contract_ids, uninstall_event_ids, charge_item_ids)
    )
    to_generate = [i for i in simulation['items'] if not i['already_generated']]
    if exact_selection:
        selected_contracts = set(contract_ids or [])
        to_generate = [
            item for item in to_generate
            if item['contract_id'] in selected_contracts
        ]

    deferred_item_ids = {
        item['event_id'] for item in simulation['uninstall_events'] if item['deferred']
    }
    # Recupera estados terminais gravados pela implementação antiga, mesmo que
    # esses eventos não sejam selecionados para faturar nesta rodada.
    for event_id in deferred_item_ids:
        deferred_event = db.get(UninstallEvent, event_id)
        if deferred_event and deferred_event.status == 'skipped':
            deferred_event.status = 'pending'
            deferred_event.processed_at = None

    selected_uninstall_items = [
        item for item in simulation['uninstall_events'] if not item['deferred']
    ]
    if exact_selection:
        selected = set(uninstall_event_ids or [])
        selected_uninstall_items = [
            item for item in selected_uninstall_items if item['event_id'] in selected
        ]

    selected_charge_items = list(simulation['charge_items'])
    if exact_selection:
        selected = set(charge_item_ids or [])
        selected_charge_items = [
            item for item in selected_charge_items if item['item_id'] in selected
        ]

    created_ids = []
    for item in to_generate:
        contract = db.get(Contract, item['contract_id'])
        if not contract:
            continue
        plan = db.get(Plan, contract.plan_id)
        if not plan:
            continue

        billing_amount = _quantize_amount(item['billing_amount'])
        first_charges = item.get('first_month_charges', [])

        if first_charges:
            # Cobrança combinada: mensalidade + serviços do primeiro mês
            service_total = sum(Decimal(str(c['amount'])) for c in first_charges)
            combined_amount = billing_amount + service_total

            service_parts = ' | '.join(
                f'{c["title"]}: R$ {c["amount"]:.2f}' for c in first_charges
            )
            if item['is_prorata']:
                plan_label = f'{plan_title(plan)} pró-rata {item["prorated_days"]} dias'
            else:
                plan_label = plan_title(plan)

            title = f'1ª cobrança — {plan_label}'
            notes = (
                f'Mensalidade ({plan.name}): R$ {float(billing_amount):.2f} | '
                + service_parts
                + f' | Total: R$ {float(combined_amount):.2f}'
            )
            billing_type = 'primeira_mensalidade'

            billing = Billing(
                contract_id=contract.id,
                client_id=contract.client_id,
                payer_client_id=contract_payer_client_id(db, contract),
                vehicle_id=getattr(contract, 'vehicle_id', None),
                tracker_id=getattr(contract, 'tracker_id', None),
                amount=combined_amount,
                due_date=item['due_date'],
                status=BillingStatus.PENDING if item['due_date'] >= hoje() else BillingStatus.OVERDUE,
                period_label=item['period_label'],
                payment_method=contract.payment_method,
                notes=notes,
                title=title,
                billing_type=billing_type,
            )
            db.add(billing)
            db.flush()
            created_ids.append(billing.id)

            # Registra cada serviço no título combinado. A emissão apenas o marca
            # como faturado; a conclusão é derivada da baixa do pagamento.
            for charge in first_charges:
                charge_obj = db.get(ClientChargeItem, charge['item_id'])
                if charge_obj:
                    associate_billing_charge_item(db, billing, charge_obj, charge['amount'])
                    charge_obj.active = False
                    charge_obj.completed_at = None
                    charge_obj.status = 'faturado'

        else:
            # Cobrança normal (mensalidade ou pró-rata sem serviços embutidos)
            if item['is_prorata']:
                title = f'{plan_title(plan)} — pró-rata {item["prorated_days"]} dias'
                notes = (
                    f'Pró-rata: {item["prorated_days"]} de {item["days_in_month"]} dias'
                    f' — {item["period_label"]}'
                )
                billing_type = 'prorata'
            else:
                title = plan_title(plan)
                notes = f'Fechamento — {item["period_label"]}'
                billing_type = 'recorrente'

            billing = Billing(
                contract_id=contract.id,
                client_id=contract.client_id,
                payer_client_id=contract_payer_client_id(db, contract),
                vehicle_id=getattr(contract, 'vehicle_id', None),
                tracker_id=getattr(contract, 'tracker_id', None),
                amount=billing_amount,
                due_date=item['due_date'],
                status=BillingStatus.PENDING if item['due_date'] >= hoje() else BillingStatus.OVERDUE,
                period_label=item['period_label'],
                payment_method=contract.payment_method,
                notes=notes,
                title=title,
                billing_type=billing_type,
            )
            db.add(billing)
            db.flush()
            created_ids.append(billing.id)

    # ── Boleto único por cliente (boleto_format='unico' no cadastro) ────────
    # Junta as MENSALIDADES normais recém-criadas do mesmo cliente em UMA
    # cobrança só (1 boleto = 1 tarifa bancária/mês, como o campo do cadastro
    # promete). Pró-rata e 1ª cobrança ficam de fora (semântica própria).
    # As individuais são canceladas com referência cruzada — e continuam
    # contando para a idempotência (_has_existing_billing ignora o status).
    consolidated_ids: list[int] = []
    _por_pagador: dict[int, list[Billing]] = defaultdict(list)
    for bid in created_ids:
        b = db.get(Billing, bid)
        if b and b.billing_type == 'recorrente':
            _por_pagador[b.payer_client_id or b.client_id].append(b)

    for payer_id, grupo in _por_pagador.items():
        if len(grupo) < 2:
            continue
        cliente = db.get(Client, payer_id)
        # Só consolida com a opção EXPLÍCITA no cadastro (campo vazio = individual)
        if not cliente or cliente.boleto_format != 'unico':
            continue

        total = sum(Decimal(str(b.amount)) for b in grupo)
        venc = max(b.due_date for b in grupo)
        period = grupo[0].period_label

        def _placa(b: Billing) -> str:
            v = db.get(Vehicle, b.vehicle_id) if b.vehicle_id else None
            return v.plate if v and not v.is_deleted else (b.title or f'#{b.id}')

        detalhes = ' | '.join(f'{_placa(b)}: R$ {float(b.amount):.2f}' for b in grupo)
        owner_ids = {billing.client_id for billing in grupo}
        unico = Billing(
            # Se há vários clientes atendidos pelo mesmo interveniente, o título
            # consolidado não pertence exclusivamente a nenhum deles.
            client_id=(next(iter(owner_ids)) if len(owner_ids) == 1 else payer_id),
            payer_client_id=payer_id,
            billing_type='recorrente',
            title=f'Mensalidades — {len(grupo)} veículos (boleto único)',
            amount=total,
            due_date=venc,
            status=BillingStatus.PENDING if venc >= hoje() else BillingStatus.OVERDUE,
            period_label=period,
            payment_method=grupo[0].payment_method,
            notes=f'Boleto único ({period}): {detalhes}',
        )
        db.add(unico)
        db.flush()
        for b in grupo:
            b.status = BillingStatus.CANCELED
            marker = f'Consolidada no boleto único #{unico.id}.'
            b.notes = f'{b.notes} | {marker}' if b.notes else marker
            created_ids.remove(b.id)
        created_ids.append(unico.id)
        consolidated_ids.append(unico.id)

    # Processa SOMENTE os eventos que pertenciam à simulação e foram enviados
    # pelo cliente da API. Isso fecha o TOCTOU em que uma taxa criada depois da
    # prévia entrava silenciosamente no fechamento.
    uninstall_billing_ids: list[int] = []
    deferred_events = len(deferred_item_ids)
    processed_events = 0
    allowed_event_ids = {item['event_id'] for item in selected_uninstall_items}
    uninstall_events = [
        event for event in _pending_uninstall_events_for_month(
            db, reference_month, filter_type, client_id,
        )
        if event.id in allowed_event_ids
    ]
    now_utc = datetime.now(timezone.utc)

    events_by_payer: dict[int, list[UninstallEvent]] = defaultdict(list)
    for event in uninstall_events:
        events_by_payer[_uninstall_event_payer_client_id(db, event)].append(event)

    for event_payer_id, events in events_by_payer.items():
        event_amounts = {
            event.id: uninstall_fee_for_event(db, event)[0]
            for event in events
        }
        total_fee = sum(event_amounts.values(), Decimal('0.00'))
        if total_fee < MIN_BILLING_AMOUNT:
            # Recupera inclusive os antigos ``skipped``. O grupo permanece
            # pendente e poderá ser somado a eventos de competências futuras.
            for event in events:
                event.status = 'pending'
                event.processed_at = None
            deferred_events += len(events)
            continue

        due_date = max(_due_date_for_uninstall_event(event, db) for event in events)
        # Só associa a cobrança agregada a um contrato quando TODOS os eventos
        # pertencem explicitamente ao mesmo contrato. Misturar evento sem
        # contrato com evento contratado e escolher o único ID conhecido
        # produziria uma associação contábil enganosa.
        contract_ids_in_group = {event.contract_id for event in events}
        owner_ids_in_group = {event.client_id for event in events}
        single_event = events[0] if len(events) == 1 else None
        if single_event is not None:
            _, fee_title = uninstall_fee_for_event(db, single_event)
        else:
            fee_title = f'Taxas de desinstalação agrupadas ({len(events)} eventos)'

        detail_parts = []
        for event in events:
            vehicle = db.get(Vehicle, event.vehicle_id)
            detail_parts.append(
                f'#{event.id} {vehicle.plate if vehicle else "veículo removido"} '
                f'{event.uninstall_date.strftime("%d/%m/%Y")}: '
                f'R$ {float(event_amounts[event.id]):.2f}'
            )
        fee_billing = Billing(
            contract_id=(
                next(iter(contract_ids_in_group))
                if len(contract_ids_in_group) == 1 and None not in contract_ids_in_group
                else None
            ),
            client_id=(
                next(iter(owner_ids_in_group))
                if len(owner_ids_in_group) == 1 else event_payer_id
            ),
            payer_client_id=event_payer_id,
            vehicle_id=(single_event.vehicle_id if single_event else None),
            tracker_id=(single_event.tracker_id if single_event else None),
            title=fee_title,
            billing_type='taxa_desinstalacao',
            amount=total_fee,
            due_date=due_date,
            status=BillingStatus.PENDING if due_date >= hoje() else BillingStatus.OVERDUE,
            period_label=due_date.strftime('%m/%Y'),
            notes='Taxas processadas no fechamento: ' + ' | '.join(detail_parts),
        )
        db.add(fee_billing)
        db.flush()
        for event in events:
            event.status = 'processed'
            event.billing_id = fee_billing.id
            event.processed_at = now_utc
            processed_events += 1
        uninstall_billing_ids.append(fee_billing.id)

    # Gera billings para serviços/cobranças avulsas pendentes (os não embutidos)
    services_generated = 0
    service_billing_ids: list[int] = []
    for charge_item_dict in selected_charge_items:
        item_obj = db.get(ClientChargeItem, charge_item_dict['item_id'])
        if item_obj:
            new_billings = generate_item_billings(db, item_obj, commit=False)
            for b in new_billings:
                service_billing_ids.append(b.id)
            services_generated += len(new_billings)

    # Único commit do fechamento: até aqui nada foi confirmado, então uma falha
    # em qualquer etapa acima desfaz o fechamento inteiro em vez de deixar
    # metade das cobranças gravadas.
    db.commit()
    refresh_overdue_statuses(db)

    # Mensalidades: soma o valor real criado (combined ou normal)
    total_mensalidades = round(
        sum(float(db.get(Billing, bid).amount) for bid in created_ids), 2
    )
    total_uninstall_amount = round(
        sum(float(db.get(Billing, bid).amount) for bid in uninstall_billing_ids), 2
    )
    total_services_amount = round(
        sum(float(db.get(Billing, bid).amount) for bid in service_billing_ids), 2
    )

    return {
        'reference_month': simulation['reference_month'],
        'generated': len(created_ids),
        'billing_ids': created_ids,
        'consolidated_unico': len(consolidated_ids),
        'total_amount': total_mensalidades,
        'uninstall_fees_generated': len(uninstall_billing_ids),
        'uninstall_events_processed': processed_events,
        'uninstall_fees_deferred': deferred_events,
        # Compatibilidade com consumidores antigos: eventos não são mais
        # descartados, portanto o total de ignorados é sempre zero.
        'uninstall_fees_skipped': 0,
        'uninstall_billing_ids': uninstall_billing_ids,
        'services_generated': services_generated,
        'service_billing_ids': service_billing_ids,
        'total_services_amount': total_services_amount,
        'grand_total': round(total_mensalidades + total_uninstall_amount + total_services_amount, 2),
    }



# ---------------------------------------------------------------------------
# Relatório de simulação de fechamento (formato do sistema antigo)
# ---------------------------------------------------------------------------
# Texto monoespaçado, agrupado por INTERVENIENTE (quem paga o boleto), com um
# bloco por veículo e uma linha por rastreador — um veículo pode ter mais de um
# equipamento, e cada um tem a própria mensalidade.

_LARGURA = 96          # colunas do relatório
_SEP = '=' * _LARGURA
_EMPRESA = 'MASTERSAT COMERCIO E SERVIÇOS DE RASTREAMENTO LTDA'


def _v(valor: float) -> str:
    """Valor no padrão brasileiro, sem símbolo (o relatório é monoespaçado)."""
    return f'{valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _d(valor: date | None) -> str:
    return valor.strftime('%d/%m/%Y') if valor else '//'


def _par(esquerda: str, direita: str, largura: int = _LARGURA) -> str:
    """Duas colunas na mesma linha: uma à esquerda, outra à direita."""
    espaco = max(1, largura - len(esquerda) - len(direita))
    return f'{esquerda}{" " * espaco}{direita}'


def _total(rotulo: str, valor: float, recuo: int = 40) -> str:
    """
    Linha de total: rótulo recuado e valor à direita.

    Os 12 caracteres finais alinham na MESMA coluna do valor das linhas de
    rastreador (8 + 74 + 12 = 94), senão a coluna de valores fica em degrau.
    """
    corpo = f'{rotulo:<42}{_v(valor):>12}'
    return ' ' * recuo + corpo


def _chave_veiculo(item: dict) -> object:
    """Sem veículo, cada contrato é seu próprio grupo."""
    return item.get('vehicle_id') or f'c{item["contract_id"]}'


def _plural(n: int, singular: str, plural: str) -> str:
    return f'{n} {singular if n == 1 else plural}'


def _no_mes(valor: date | None, mes_ref: str) -> bool:
    """A data cai no mês de referência ('MM/AAAA')?"""
    if not valor or '/' not in (mes_ref or ''):
        return False
    mes, _, ano = mes_ref.partition('/')
    return valor.month == int(mes) and valor.year == int(ano)


def _contagens(itens: list[dict], desinstalacoes: list[dict], mes_ref: str) -> dict[str, int]:
    """Veículos, equipamentos, instalações e desinstalações de um conjunto."""
    return {
        'veiculos': len({_chave_veiculo(i) for i in itens}),
        # Um veículo pode ter mais de um rastreador — por isso a contagem é
        # separada da de veículos (ex.: ACQUE, 1 veículo com 2 equipamentos).
        'equipamentos': len({i['tracker_imei'] for i in itens if i.get('tracker_imei')}),
        'instalacoes': sum(1 for i in itens if _no_mes(i.get('tracker_install_date'), mes_ref)),
        'desinstalacoes': len(desinstalacoes),
    }


_MESES_EXT = ['', 'JANEIRO', 'FEVEREIRO', 'MARÇO', 'ABRIL', 'MAIO', 'JUNHO',
              'JULHO', 'AGOSTO', 'SETEMBRO', 'OUTUBRO', 'NOVEMBRO', 'DEZEMBRO']


def _mes_extenso(mes_ref: str) -> str:
    """'08/2026' → 'AGOSTO/2026'; devolve o original se não bater o formato."""
    mm, _, aaaa = (mes_ref or '').partition('/')
    if aaaa and mm.isdigit() and 1 <= int(mm) <= 12:
        return f'{_MESES_EXT[int(mm)]}/{aaaa}'
    return mes_ref or ''


def _totais_financeiros(itens: list[dict], simulation: dict) -> dict[str, float]:
    """Somatórios do fechamento, por natureza — a mesma conta no topo e no rodapé."""
    mensalidades = sum(float(i.get('billing_amount') or 0) for i in itens)
    produtos = sum(
        float(p.get('amount') or 0)
        for i in itens for p in (i.get('first_month_charges') or [])
    )
    taxas = float(simulation.get('total_uninstall_fees') or 0)
    servicos = float(simulation.get('total_services') or 0)
    return {
        'mensalidades': mensalidades, 'produtos': produtos,
        'taxas': taxas, 'servicos': servicos,
        'geral': mensalidades + produtos + taxas + servicos,
    }


def _painel_totais(contagem: dict[str, int], fin: dict[str, float]) -> list[str]:
    """
    Painel de totais no topo do relatório — a quantidade de tudo antes do
    detalhamento por cliente (pedido de 08/08/2026). Tabela de 5 colunas em
    caixa, no monoespaçado do relatório.
    """
    colunas = [
        ('Veículos', str(contagem['veiculos'])),
        ('Rastreadores', str(contagem['equipamentos'])),
        ('Instalações', str(contagem['instalacoes'])),
        ('Desinstalações', str(contagem['desinstalacoes'])),
        ('Total Geral', f'R$ {_v(fin["geral"])}'),
    ]
    w = 18  # 5 células de 18 + 6 bordas = 96 = _LARGURA
    borda = '+' + '+'.join('-' * w for _ in colunas) + '+'

    def _linha(celulas: list[str]) -> str:
        return '|' + '|'.join(c.center(w) for c in celulas) + '|'

    return [
        borda,
        _linha([rot for rot, _ in colunas]),
        borda,
        _linha([val for _, val in colunas]),
        borda,
    ]


def montar_linhas_simulacao(simulation: dict) -> list[str]:
    """
    Monta o relatório como lista de linhas de texto.

    Separado da geração do PDF para poder ser testado sem abrir o arquivo.
    """
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    mes_ref = simulation.get('reference_month', '')
    titulo = f'PRÉVIA DE FECHAMENTO — {_mes_extenso(mes_ref)}'.strip(' —')
    linhas: list[str] = [_SEP, titulo.center(_LARGURA), _SEP]

    itens = simulation.get('items') or []
    if not itens:
        linhas += ['', 'Nenhum contrato a faturar no período.']
        return linhas

    # 1) agrupa por interveniente  2) dentro dele, por veículo
    por_interveniente: dict[str, list[dict]] = defaultdict(list)
    for item in itens:
        por_interveniente[item.get('interveniente_nome') or item['client_name']].append(item)

    # As desinstalações vêm por cliente; o relatório agrupa por interveniente.
    # Este mapa leva uma à outra para a contagem cair no bloco certo.
    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }
    desinst_por_grupo: dict[str, list[dict]] = defaultdict(list)
    for ev in simulation.get('uninstall_events') or []:
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        desinst_por_grupo[grupo].append(ev)

    fin = _totais_financeiros(itens, simulation)

    # ── Painel de totais no topo (a quantidade de tudo antes dos clientes) ──
    linhas.append(_par(f'Mês de referência: {mes_ref}', f'Gerado em {agora}'))
    linhas.append('')
    linhas += _painel_totais(
        _contagens(itens, simulation.get('uninstall_events') or [], mes_ref), fin)
    linhas.append('')

    for interveniente, do_grupo in sorted(por_interveniente.items()):
        por_veiculo: dict[object, list[dict]] = defaultdict(list)
        for item in do_grupo:
            por_veiculo[_chave_veiculo(item)].append(item)

        venc = do_grupo[0].get('due_date')
        mes_venc = venc.strftime('%m/%Y') if venc else ''
        qtd = _contagens(do_grupo, desinst_por_grupo.get(interveniente, []), mes_ref)

        linhas.append(f'INTERVENIENTE: {interveniente}')
        linhas.append(f'MATRIZ/FILIAL: {_EMPRESA}')
        linhas.append('')
        linhas.append(_par(
            f'MÊS REFERENTE: {mes_ref}',
            _par(f'MÊS VENCIMENTO: {mes_venc}',
                 f'QUANTIDADE VEÍCULOS: {qtd["veiculos"]}', 62),
        ))
        linhas.append(_par(
            f'QUANTIDADE EQUIPAMENTOS: {qtd["equipamentos"]}',
            _par(f'INSTALAÇÕES NO MÊS: {qtd["instalacoes"]}',
                 f'DESINSTALAÇÕES NO MÊS: {qtd["desinstalacoes"]}', 62),
        ))
        linhas.append('')

        total_grupo = 0.0
        for contratos in por_veiculo.values():
            primeiro = contratos[0]
            venc_v = primeiro.get('due_date')
            dia = primeiro.get('billing_day') or (venc_v.day if venc_v else '')

            linhas.append(f'CLIENTE: {primeiro["client_name"]}')
            linhas.append(_par(
                f'PLACA: {primeiro.get("vehicle_plate") or "—"}',
                f'TIPO VEÍCULO: {(primeiro.get("vehicle_type") or "—").upper()}', 78))
            linhas.append(_par(
                f'DATA CADASTRO: {_d(primeiro.get("vehicle_created_at"))}',
                f'RASTREADOR: {primeiro.get("tracker_imei") or "—"}', 78))
            linhas.append(_par(
                f'DATA CONTRATO: {_d(primeiro.get("contract_start_date"))}',
                f'VENCIMENTO: {dia}[{_d(venc_v)}]', 78))
            linhas.append('')

            total_veiculo = 0.0
            for c in contratos:
                mensalidade = float(c.get('billing_amount') or 0)
                instalacao = _d(c.get('tracker_install_date'))
                rotulo = f'RASTREADOR: DATA INSTALAÇÃO [{instalacao}] - MENSALIDADE {_v(mensalidade)}:'
                linhas.append(' ' * 8 + f'{rotulo:<74}{_v(mensalidade):>12}'.rstrip())

                # Serviços/produtos embutidos na primeira cobrança
                produtos = c.get('first_month_charges') or []
                soma_produtos = 0.0
                for prod in produtos:
                    val = float(prod.get('amount') or 0)
                    soma_produtos += val
                    desc = f'PRODUTO - {str(prod.get("title") or "").upper()}:'
                    linhas.append(' ' * 8 + f'{desc:<74}{_v(val):>12}'.rstrip())
                if produtos:
                    linhas.append(_total('SOMA PRODUTOS:', soma_produtos))

                linhas.append(_total('TOTAL RASTREADOR:', mensalidade))
                linhas.append('')
                total_veiculo += mensalidade + soma_produtos

            linhas.append(_total('TOTAL VEÍCULO:', total_veiculo))
            linhas.append('')
            total_grupo += total_veiculo

        # Sem impostos configurados, os dois totais são iguais — as duas linhas
        # existem porque o relatório de referência as traz.
        linhas.append(_total('TOTAL BOLETO S/ IMPOSTOS:', total_grupo))
        linhas.append(_total('TOTAL BOLETO C/ IMPOSTOS:', total_grupo))
        linhas.append(_SEP)

    linhas += _movimentacao_do_mes(itens, simulation, por_interveniente, mes_ref)
    linhas += _resumo_geral(simulation, itens, por_interveniente, mes_ref, fin)
    return linhas


def _movimentacao_do_mes(itens: list[dict], simulation: dict,
                         por_interveniente: dict[str, list[dict]],
                         mes_ref: str) -> list[str]:
    """
    Instalações e desinstalações do período, uma a uma.

    Pedido da reunião de 07/08/2026: o resumo diz quantas foram; aqui se vê
    quais — por interveniente e por cliente, com data, placa e equipamento.
    """
    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }

    instalacoes: dict[str, list[dict]] = defaultdict(list)
    for item in itens:
        if _no_mes(item.get('tracker_install_date'), mes_ref):
            instalacoes[item.get('interveniente_nome') or item['client_name']].append(item)

    desinstalacoes: dict[str, list[dict]] = defaultdict(list)
    for ev in simulation.get('uninstall_events') or []:
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        desinstalacoes[grupo].append(ev)

    if not instalacoes and not desinstalacoes:
        return []

    linhas = ['', f'MOVIMENTAÇÃO DO PERÍODO — {mes_ref}'.center(_LARGURA), _SEP]

    for grupo in sorted(set(instalacoes) | set(desinstalacoes)):
        linhas.append(f'INTERVENIENTE: {grupo}')

        for item in sorted(instalacoes.get(grupo, []),
                           key=lambda i: (i.get('tracker_install_date') or date.min,
                                          i.get('vehicle_plate') or '')):
            linhas.append(_par(
                f'    INSTALAÇÃO  {_d(item.get("tracker_install_date"))}  '
                f'{(item.get("vehicle_plate") or "SEM PLACA"):<10} '
                f'RASTREADOR {item.get("tracker_imei") or "—"}',
                item['client_name'][:34],
            ))

        for ev in sorted(desinstalacoes.get(grupo, []),
                         key=lambda e: (e.get('uninstall_date') or date.min,
                                        e.get('vehicle_plate') or '')):
            taxa = float(ev.get('fee_amount') or 0)
            # O valor pequeno continua devido e será acumulado; o relatório não
            # pode apresentá-lo como isento ou descartado.
            sufixo = (
                f'TAXA {_v(taxa)} (AGUARDANDO ACUMULAÇÃO)'
                if ev.get('deferred') else f'TAXA {_v(taxa)}'
            )
            linhas.append(_par(
                f'    DESINSTALAÇÃO  {_d(ev.get("uninstall_date"))}  '
                f'{(ev.get("vehicle_plate") or "SEM PLACA"):<10} {sufixo}',
                (ev.get('client_name') or '')[:34],
            ))

        linhas.append('    Subtotal: ' + ' · '.join((
            _plural(len(instalacoes.get(grupo, [])), 'instalação', 'instalações'),
            _plural(len(desinstalacoes.get(grupo, [])), 'desinstalação', 'desinstalações'),
        )))
        linhas.append('')

    linhas.append(_SEP)
    return linhas


def _resumo_geral(simulation: dict, itens: list[dict],
                  por_interveniente: dict[str, list[dict]], mes_ref: str,
                  fin: dict[str, float]) -> list[str]:
    """Fecha o relatório com o consolidado do período (mesma conta do topo)."""
    desinstalacoes = simulation.get('uninstall_events') or []
    qtd = _contagens(itens, desinstalacoes, mes_ref)

    linhas = ['', f'RESUMO DO FECHAMENTO — {mes_ref}'.center(_LARGURA), _SEP]
    linhas.append(_par(
        f'INTERVENIENTES: {len(por_interveniente)}',
        _par(f'VEÍCULOS: {qtd["veiculos"]}',
             f'EQUIPAMENTOS: {qtd["equipamentos"]}', 62),
    ))
    linhas.append(_par(
        f'INSTALAÇÕES NO MÊS: {qtd["instalacoes"]}',
        _par(f'DESINSTALAÇÕES NO MÊS: {qtd["desinstalacoes"]}',
             f'CONTRATOS: {len(itens)}', 62),
    ))

    # Cobranças já geradas continuam listadas no detalhamento; sem a linha, o
    # total do resumo pareceria não bater com o que o fechamento vai gerar.
    ja_geradas = int(simulation.get('already_generated') or 0)
    if ja_geradas:
        linhas.append(f'CONTRATOS JÁ FATURADOS NO PERÍODO (inclusos acima): {ja_geradas}')

    linhas.append('')
    linhas.append(_total('TOTAL MENSALIDADES:', fin['mensalidades']))
    if fin['produtos']:
        linhas.append(_total('TOTAL PRODUTOS/SERVIÇOS NA 1ª COBRANÇA:', fin['produtos']))
    if fin['taxas']:
        linhas.append(_total('TOTAL TAXAS DE DESINSTALAÇÃO:', fin['taxas']))
    if fin['servicos']:
        linhas.append(_total('TOTAL SERVIÇOS AVULSOS:', fin['servicos']))
    linhas.append(_total('TOTAL GERAL:', fin['geral']))
    linhas.append(_SEP)
    return linhas


def generate_closure_xlsx(simulation: dict) -> BytesIO:
    """
    Simulação de fechamento em Excel (.xlsx), com 3 abas:
      Resumo        — período, painel de totais e os somatórios financeiros
      Contratos     — uma linha por contrato (interveniente, cliente, veículo…)
      Movimentação  — instalações e desinstalações do período, uma por linha

    Ao contrário do PDF (texto monoespaçado), aqui os dados vão em colunas para
    a operação filtrar, ordenar e somar na planilha.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    mes_ref = simulation.get('reference_month', '')
    itens = simulation.get('items') or []
    desinst = simulation.get('uninstall_events') or []
    fin = _totais_financeiros(itens, simulation)
    cont = _contagens(itens, desinst, mes_ref)
    agora = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    interveniente_do_cliente = {
        i.get('client_id'): (i.get('interveniente_nome') or i['client_name']) for i in itens
    }

    MONEY = '"R$" #,##0.00'
    DATA = 'DD/MM/YYYY'
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1F2A44')
    bold = Font(bold=True)

    def _cabecalho(ws, colunas: list[str]) -> None:
        for ci, nome in enumerate(colunas, 1):
            cell = ws.cell(1, ci, nome)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.freeze_panes = 'A2'

    wb = Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumo'
    ws['A1'] = 'PRÉVIA DE FECHAMENTO'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Mês de referência: {mes_ref}'
    ws['A3'] = f'Gerado em: {agora}'

    ws['A5'] = 'Totais do período'
    ws['A5'].font = bold
    linha = 6
    for rotulo, valor in (
        ('Veículos', cont['veiculos']),
        ('Rastreadores', cont['equipamentos']),
        ('Instalações', cont['instalacoes']),
        ('Desinstalações', cont['desinstalacoes']),
    ):
        ws.cell(linha, 1, rotulo).font = bold
        ws.cell(linha, 2, valor)
        linha += 1

    linha += 1
    ws.cell(linha, 1, 'Totais financeiros').font = bold
    linha += 1
    for rotulo, valor in (
        ('Total mensalidades', fin['mensalidades']),
        ('Total produtos/serviços (1ª cobrança)', fin['produtos']),
        ('Total taxas de desinstalação', fin['taxas']),
        ('Total serviços avulsos', fin['servicos']),
        ('TOTAL GERAL', fin['geral']),
    ):
        c1 = ws.cell(linha, 1, rotulo)
        c2 = ws.cell(linha, 2, float(valor))
        c2.number_format = MONEY
        if rotulo.startswith('TOTAL GERAL'):
            c1.font = bold
            c2.font = bold
        linha += 1
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    # ── Aba Contratos ───────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Contratos')
    colunas = [
        'Interveniente', 'Cliente', 'Placa', 'Tipo veículo', 'Rastreador (IMEI)',
        'Data instalação', 'Data contrato', 'Plano', 'Mensalidade',
        'Produtos 1ª cobrança', 'Total', 'Vencimento', 'Período', 'Já faturado',
    ]
    _cabecalho(ws2, colunas)
    rr = 2
    for it in itens:
        produtos = sum(float(p.get('amount') or 0) for p in (it.get('first_month_charges') or []))
        valores = [
            it.get('interveniente_nome') or it.get('client_name'),
            it.get('client_name'),
            it.get('vehicle_plate') or '',
            (it.get('vehicle_type') or '').upper(),
            it.get('tracker_imei') or '',
            it.get('tracker_install_date'),
            it.get('contract_start_date'),
            it.get('plan_name'),
            float(it.get('billing_amount') or 0),
            produtos,
            float(it.get('total_first_billing') or 0),
            it.get('due_date'),
            it.get('period_label') or '',
            'Sim' if it.get('already_generated') else 'Não',
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws2.cell(rr, ci, val)
            if ci in (6, 7, 12):
                cell.number_format = DATA
            elif ci in (9, 10, 11):
                cell.number_format = MONEY
        rr += 1
    for ci, largura in enumerate([28, 28, 10, 14, 20, 15, 15, 20, 14, 18, 14, 13, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = largura

    # ── Aba Movimentação ────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Movimentação')
    _cabecalho(ws3, ['Tipo', 'Data', 'Interveniente', 'Cliente', 'Placa', 'Rastreador', 'Taxa', 'Status'])
    rr = 2
    for it in itens:
        if _no_mes(it.get('tracker_install_date'), mes_ref):
            valores = [
                'Instalação', it.get('tracker_install_date'),
                it.get('interveniente_nome') or it.get('client_name'),
                it.get('client_name'), it.get('vehicle_plate') or '',
                it.get('tracker_imei') or '', None, '',
            ]
            for ci, val in enumerate(valores, 1):
                cell = ws3.cell(rr, ci, val)
                if ci == 2:
                    cell.number_format = DATA
            rr += 1
    for ev in desinst:
        taxa = float(ev.get('fee_amount') or 0)
        grupo = (
            ev.get('payer_name')
            or interveniente_do_cliente.get(ev.get('client_id'))
            or ev.get('client_name') or ''
        )
        valores = [
            'Desinstalação', ev.get('uninstall_date'), grupo,
            ev.get('client_name') or '', ev.get('vehicle_plate') or '', '', taxa,
            ('Aguardando acumulação' if ev.get('deferred') else 'A faturar'),
        ]
        for ci, val in enumerate(valores, 1):
            cell = ws3.cell(rr, ci, val)
            if ci == 2:
                cell.number_format = DATA
            elif ci == 7 and val is not None:
                cell.number_format = MONEY
        rr += 1
    for ci, largura in enumerate([14, 12, 28, 28, 10, 20, 12, 24], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = largura

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_closure_pdf(simulation: dict) -> BytesIO:
    """Simulação de fechamento em PDF, no formato monoespaçado do SGR."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdfcanvas

    linhas = montar_linhas_simulacao(simulation)

    buffer = BytesIO()
    c = pdfcanvas.Canvas(buffer, pagesize=A4)
    c.setTitle('Simulação de Fechamento — MasterSat')
    largura, altura = A4
    margem = 10 * mm
    topo = altura - margem
    passo = 3.3 * mm
    fonte = 6.5

    y = topo
    pagina = 1
    for linha in linhas:
        if y < margem + 8 * mm:
            c.setFont('Courier', 6)
            c.drawCentredString(largura / 2, margem, f'Página {pagina}')
            c.showPage()
            pagina += 1
            y = topo
        c.setFont('Courier', fonte)
        c.drawString(margem, y, linha)
        y -= passo

    c.setFont('Courier', 6)
    c.drawCentredString(largura / 2, margem, f'Página {pagina}')
    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer
