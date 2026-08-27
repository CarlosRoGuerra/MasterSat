"""
Exportação de dados em CSV e Excel.

Endpoints:
  GET /exports/clients    → Lista de clientes
  GET /exports/vehicles   → Lista de veículos
  GET /exports/trackers   → Lista de rastreadores
  GET /exports/billings   → Cobranças com filtros
  GET /exports/delinquents → Relatório de inadimplentes
"""
from __future__ import annotations

import csv
import io
from datetime import date

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.core.limiter import limiter

from app.api.deps import require_roles
from app.db.session import get_db
from app.models.billing import Billing
from app.models.client import Client
from app.models.contract import Contract
from app.models.enums import BillingStatus, UserRole
from app.models.plan import Plan
from app.models.tracker import Tracker
from app.models.vehicle import Vehicle

router = APIRouter()

VIEW_ROLES = (UserRole.ADMIN, UserRole.FINANCIAL, UserRole.OPERATIONAL)


# Célula de texto começando com um destes é interpretada como fórmula por
# Excel/Sheets (CSV/Excel injection). Prefixamos com apóstrofo p/ virar texto.
_FORMULA_TRIGGERS = ('=', '+', '-', '@', '\t', '\r')


def _neutralize_formula(value):
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _csv_response(headers: list[str], rows: list[list], filename: str) -> StreamingResponse:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    writer.writerows([_neutralize_formula(c) for c in row] for row in rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue().encode('utf-8-sig')]),  # utf-8-sig = BOM para Excel abrir correto
        media_type='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _excel_response(headers: list[str], rows: list[list], filename: str) -> StreamingResponse:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Dados'

    # Cabeçalho com estilo
    header_fill = PatternFill('solid', fgColor='1E3A5F')
    header_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=_neutralize_formula(val))

    # Auto-ajusta largura das colunas
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


def _build_response(fmt: str, headers: list[str], rows: list[list], name: str) -> StreamingResponse:
    if fmt == 'xlsx':
        return _excel_response(headers, rows, f'{name}.xlsx')
    return _csv_response(headers, rows, f'{name}.csv')


# ---------------------------------------------------------------------------
# Exportar clientes
# ---------------------------------------------------------------------------

@router.get('/clients')
@limiter.limit('10/minute')
def export_clients(
    request: Request,
    fmt: str = Query(default='csv', pattern='^(csv|xlsx)$'),
    status: str | None = None,
    type: str | None = None,
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    query = db.query(Client).filter(Client.is_deleted.is_(False))
    if status:
        query = query.filter(Client.status == status)
    if type:
        query = query.filter(Client.type == type)
    clients = query.order_by(Client.name).all()

    headers = ['ID', 'Nome', 'CPF/CNPJ', 'Tipo', 'Status', 'E-mail', 'Telefone',
               'Cidade', 'UF', 'Dia Vencimento', 'Data Cadastro']
    rows = [
        [
            c.id, c.name, c.cpf_cnpj, c.type.upper(),
            c.status.value if hasattr(c.status, 'value') else str(c.status),
            c.email or '', c.phone or '', c.city or '', c.state or '',
            c.billing_day or '',
            c.created_at.strftime('%d/%m/%Y') if c.created_at else '',
        ]
        for c in clients
    ]
    return _build_response(fmt, headers, rows, 'clientes')


# ---------------------------------------------------------------------------
# Exportar veículos
# ---------------------------------------------------------------------------

@router.get('/vehicles')
def export_vehicles(
    fmt: str = Query(default='csv', pattern='^(csv|xlsx)$'),
    status: str | None = None,
    client_id: int | None = None,
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    query = db.query(Vehicle).filter(Vehicle.is_deleted.is_(False))
    if status:
        query = query.filter(Vehicle.status == status)
    if client_id:
        query = query.filter(Vehicle.client_id == client_id)
    vehicles = query.order_by(Vehicle.plate).all()

    client_map = {c.id: c.name for c in db.query(Client).filter(Client.is_deleted.is_(False)).all()}
    tracker_map: dict[int, Tracker] = {
        t.vehicle_id: t
        for t in db.query(Tracker).filter(
            Tracker.is_deleted.is_(False),
            Tracker.vehicle_id.isnot(None),
        ).all()
        if t.vehicle_id
    }

    headers = ['ID', 'Placa', 'Marca', 'Modelo', 'Ano Modelo', 'Tipo', 'Cor',
               'Chassi', 'Renavam', 'Status', 'Cliente', 'IMEI Rastreador', 'Data Cadastro']
    rows = [
        [
            v.id, v.plate, v.brand or '', v.model or '', v.model_year or '',
            v.type or '',v.color or '', v.chassis or '', v.renavam or '',
            v.status.value if hasattr(v.status, 'value') else str(v.status),
            client_map.get(v.client_id, ''),
            tracker_map[v.id].imei if v.id in tracker_map else '',
            v.created_at.strftime('%d/%m/%Y') if v.created_at else '',
        ]
        for v in vehicles
    ]
    return _build_response(fmt, headers, rows, 'veiculos')


# ---------------------------------------------------------------------------
# Exportar rastreadores
# ---------------------------------------------------------------------------

@router.get('/trackers')
def export_trackers(
    fmt: str = Query(default='csv', pattern='^(csv|xlsx)$'),
    status: str | None = None,
    client_id: int | None = None,
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    query = db.query(Tracker).filter(Tracker.is_deleted.is_(False))
    if status:
        query = query.filter(Tracker.status == status)
    if client_id:
        query = query.filter(Tracker.client_id == client_id)
    trackers = query.order_by(Tracker.imei).all()

    client_map = {c.id: c.name for c in db.query(Client).filter(Client.is_deleted.is_(False)).all()}
    vehicle_map = {v.id: v.plate for v in db.query(Vehicle).filter(Vehicle.is_deleted.is_(False)).all()}

    headers = ['ID', 'IMEI', 'Marca', 'Modelo', 'Status', 'Cliente', 'Veículo (Placa)',
               'SIM (MSISDN)', 'ICCID', 'Operadora', 'Data Instalação', 'Garantia Até']
    rows = [
        [
            t.id, t.imei, t.brand or '', t.model or '',
            t.status.value if hasattr(t.status, 'value') else str(t.status),
            client_map.get(t.client_id, '') if t.client_id else '',
            vehicle_map.get(t.vehicle_id, '') if t.vehicle_id else '',
            t.sim_number or '', t.sim_iccid or '', t.carrier or '',
            t.install_date.strftime('%d/%m/%Y') if t.install_date else '',
            t.warranty_until.strftime('%d/%m/%Y') if t.warranty_until else '',
        ]
        for t in trackers
    ]
    return _build_response(fmt, headers, rows, 'rastreadores')


# ---------------------------------------------------------------------------
# Exportar cobranças
# ---------------------------------------------------------------------------

@router.get('/billings')
def export_billings(
    fmt: str = Query(default='csv', pattern='^(csv|xlsx)$'),
    status: str | None = None,
    client_id: int | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    query = db.query(Billing).filter(Billing.is_deleted.is_(False))
    if status:
        query = query.filter(Billing.status == status)
    if client_id:
        query = query.filter(Billing.client_id == client_id)
    if date_from:
        query = query.filter(Billing.due_date >= date_from)
    if date_to:
        query = query.filter(Billing.due_date <= date_to)
    billings = query.order_by(Billing.due_date.desc()).limit(5000).all()

    client_map = {c.id: c.name for c in db.query(Client).filter(Client.is_deleted.is_(False)).all()}
    vehicle_map = {v.id: v.plate for v in db.query(Vehicle).filter(Vehicle.is_deleted.is_(False)).all()}

    headers = ['ID', 'Cliente', 'Veículo', 'Título', 'Tipo', 'Vencimento', 'Valor',
               'Valor Pago', 'Status', 'Data Pagamento', 'Forma Pagamento',
               'Parcela', 'Total Parcelas', 'Período', 'Nº Recibo']
    rows = [
        [
            b.id,
            client_map.get(b.client_id, ''),
            vehicle_map.get(b.vehicle_id, '') if b.vehicle_id else '',
            b.title or '',
            b.billing_type or '',
            b.due_date.strftime('%d/%m/%Y') if b.due_date else '',
            float(b.amount),
            float(b.paid_amount) if b.paid_amount else 0.0,
            b.status.value if hasattr(b.status, 'value') else str(b.status),
            b.payment_date.strftime('%d/%m/%Y') if b.payment_date else '',
            b.payment_method or '',
            b.installment_number or 1,
            b.installment_total or 1,
            b.period_label or '',
            b.receipt_number or '',
        ]
        for b in billings
    ]
    return _build_response(fmt, headers, rows, 'cobrancas')


# ---------------------------------------------------------------------------
# Relatório de cobranças por período (pagas / em aberto / vencidas)
# ---------------------------------------------------------------------------

_SITUACAO_LABEL = {
    'paga': 'Paga', 'pendente': 'Em aberto', 'vencida': 'Vencida', 'cancelada': 'Cancelada',
}


def _billings_report_pdf(linhas, situacao: str, periodo_por: str,
                         date_from: date | None, date_to: date | None) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title='Relatório de Cobranças — MasterSat')
    styles = getSampleStyleSheet()
    titulo = {
        'paga': 'Cobranças Pagas', 'pendente': 'Cobranças em Aberto',
        'vencida': 'Cobranças Vencidas', 'cancelada': 'Cobranças Canceladas',
    }.get(situacao, 'Cobranças')
    por = 'pagamento' if periodo_por == 'pagamento' else 'vencimento'
    de = date_from.strftime('%d/%m/%Y') if date_from else '—'
    ate = date_to.strftime('%d/%m/%Y') if date_to else '—'

    elems = [
        Paragraph(f'MASTERSAT — Relatório de {titulo}', styles['Title']),
        Paragraph(f'Emitido em {date.today().strftime("%d/%m/%Y")}', styles['Normal']),
        Paragraph(f'<b>PERÍODO DE {de} ATÉ {ate}</b> (por data de {por})', styles['Normal']),
        Spacer(1, 8),
    ]

    header = ['Cliente', 'Título', 'Vencimento', 'Pagamento', 'Valor', 'Valor Pago', 'Situação']
    data = [header]
    total, total_pago = 0.0, 0.0
    for ln in linhas:
        total += ln['valor']
        total_pago += ln['valor_pago']
        data.append([
            ln['cliente'], ln['titulo'],
            ln['vencimento'].strftime('%d/%m/%Y') if ln['vencimento'] else '',
            ln['pagamento'].strftime('%d/%m/%Y') if ln['pagamento'] else '—',
            _brl(ln['valor']), _brl(ln['valor_pago']) if ln['valor_pago'] else '—',
            _SITUACAO_LABEL.get(ln['situacao'], ln['situacao']),
        ])
    data.append(['', '', '', 'TOTAL', _brl(total), _brl(total_pago), f'{len(linhas)} cobrança(s)'])

    table = Table(data, repeatRows=1,
                  colWidths=[70 * mm, 55 * mm, 27 * mm, 27 * mm, 30 * mm, 30 * mm, 26 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F1F5F9')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#DCFCE7')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elems.append(table)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/pdf',
        headers={'Content-Disposition': 'inline; filename="relatorio-cobrancas.pdf"'},
    )


@router.get('/billings-report')
def export_billings_report(
    fmt: str = Query(default='pdf', pattern='^(csv|xlsx|pdf)$'),
    situacao: str = Query(default='paga', pattern='^(paga|pendente|vencida|cancelada|todas)$'),
    periodo_por: str = Query(default='pagamento', pattern='^(pagamento|vencimento)$',
                             description='Qual data o período filtra'),
    date_from: date | None = None,
    date_to: date | None = None,
    client_id: int | None = None,
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    """
    Relatório de cobranças por período.

    ``periodo_por`` é o que responde "quais clientes pagaram do dia 10 ao 15":
    com 'pagamento', o período filtra a DATA DE PAGAMENTO; com 'vencimento',
    filtra o vencimento (útil para o que está em aberto).
    """
    campo = Billing.payment_date if periodo_por == 'pagamento' else Billing.due_date

    query = (
        db.query(Billing, Client.name.label('cliente'))
        .join(
            Client,
            Client.id == func.coalesce(Billing.payer_client_id, Billing.client_id),
        )
        .filter(Billing.is_deleted.is_(False), Client.is_deleted.is_(False))
    )
    if situacao != 'todas':
        query = query.filter(Billing.status == BillingStatus(situacao))
    if client_id:
        query = query.filter(
            func.coalesce(Billing.payer_client_id, Billing.client_id) == client_id
        )
    if date_from:
        query = query.filter(campo >= date_from)
    if date_to:
        query = query.filter(campo <= date_to)
    # Filtrar por pagamento sem data pagas seria sem sentido — exclui os não pagos
    if periodo_por == 'pagamento' and (date_from or date_to):
        query = query.filter(Billing.payment_date.isnot(None))

    linhas = [
        {
            'cliente': cliente,
            'titulo': b.title or b.notes or '',
            'vencimento': b.due_date,
            'pagamento': b.payment_date,
            'valor': float(b.amount or 0),
            'valor_pago': float(b.paid_amount or 0),
            'situacao': b.status.value if hasattr(b.status, 'value') else str(b.status),
            'forma': b.payment_method or '',
        }
        for b, cliente in query.order_by(campo.desc().nullslast(), Client.name).limit(5000).all()
    ]

    if fmt == 'pdf':
        return _billings_report_pdf(linhas, situacao, periodo_por, date_from, date_to)

    headers = ['Cliente', 'Título', 'Vencimento', 'Pagamento', 'Valor', 'Valor Pago',
               'Situação', 'Forma de Pagamento']
    rows = [
        [
            ln['cliente'], ln['titulo'],
            ln['vencimento'].strftime('%d/%m/%Y') if ln['vencimento'] else '',
            ln['pagamento'].strftime('%d/%m/%Y') if ln['pagamento'] else '',
            ln['valor'], ln['valor_pago'],
            _SITUACAO_LABEL.get(ln['situacao'], ln['situacao']), ln['forma'],
        ]
        for ln in linhas
    ]
    return _build_response(fmt, headers, rows, 'relatorio-cobrancas')


# ---------------------------------------------------------------------------
# Exportar relatório de inadimplentes
# ---------------------------------------------------------------------------

def _brl(valor: float) -> str:
    return f'R$ {valor:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def _delinquents_pdf(linhas, due_from: date | None, due_to: date | None) -> StreamingResponse:
    """
    Relatório de inadimplentes em PDF, pronto para imprimir.

    Uma linha por COBRANÇA vencida (não por cliente): o cliente precisa ver o
    valor, o vencimento e o número do boleto de cada uma para cobrar.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title='Relatório de Inadimplentes — MasterSat')
    styles = getSampleStyleSheet()
    elems = [
        Paragraph('MASTERSAT — Relatório de Clientes Inadimplentes', styles['Title']),
        Paragraph(f'Emitido em {date.today().strftime("%d/%m/%Y")}', styles['Normal']),
    ]
    if due_from or due_to:
        de = due_from.strftime('%d/%m/%y') if due_from else '—'
        ate = due_to.strftime('%d/%m/%y') if due_to else '—'
        elems.append(Paragraph(
            f'<b>PERÍODO DE {de} ATÉ {ate}</b> (por data de vencimento)', styles['Normal']))
    elems.append(Spacer(1, 8))

    header = ['Nome', 'Valor', 'Data de Vencimento', 'Número do Boleto', 'Dias em Atraso']
    data = [header]
    total_geral = 0.0
    hoje = date.today()
    for ln in linhas:
        valor = float(ln['valor'] or 0)
        total_geral += valor
        venc = ln['vencimento']
        data.append([
            ln['cliente'],
            _brl(valor),
            venc.strftime('%d/%m/%Y') if venc else '',
            ln['nosso_numero'] or '—',
            str((hoje - venc).days) if venc else '',
        ])
    data.append(['', '', '', 'VALOR TOTAL VENCIDO', _brl(total_geral)])

    table = Table(data, repeatRows=1, colWidths=[95 * mm, 35 * mm, 40 * mm, 45 * mm, 35 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F1F5F9')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FEE2E2')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elems.append(table)
    doc.build(elems)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type='application/pdf',
        headers={'Content-Disposition': 'inline; filename="inadimplentes.pdf"'},
    )


@router.get('/delinquents')
def export_delinquents(
    fmt: str = Query(default='csv', pattern='^(csv|xlsx|pdf)$'),
    due_from: date | None = Query(default=None, description='Vencimento a partir de'),
    due_to: date | None = Query(default=None, description='Vencimento até'),
    db: Session = Depends(get_db),
    _: object = Depends(require_roles(*VIEW_ROLES)),
):
    """
    Cobranças vencidas, uma linha por cobrança (com valor, vencimento e nosso
    número), opcionalmente restritas a um período de vencimento.
    """
    from app.models.ailos_boleto import AilosBoleto

    query = (
        db.query(Billing, Client.name.label('cliente'), AilosBoleto.nosso_numero)
        .join(
            Client,
            Client.id == func.coalesce(Billing.payer_client_id, Billing.client_id),
        )
        .outerjoin(AilosBoleto, AilosBoleto.billing_id == Billing.id)
        .filter(
            Client.is_deleted.is_(False),
            Billing.is_deleted.is_(False),
            Billing.status == BillingStatus.OVERDUE,
        )
    )
    if due_from:
        query = query.filter(Billing.due_date >= due_from)
    if due_to:
        query = query.filter(Billing.due_date <= due_to)

    linhas = [
        {
            'cliente': cliente,
            'titulo': billing.title or '',
            'valor': float(billing.amount or 0),
            'vencimento': billing.due_date,
            'nosso_numero': nosso_numero,
        }
        for billing, cliente, nosso_numero in
        query.order_by(Client.name, Billing.due_date).all()
    ]

    if fmt == 'pdf':
        return _delinquents_pdf(linhas, due_from, due_to)

    headers = ['Nome', 'Título', 'Valor', 'Data de Vencimento', 'Número do Boleto', 'Dias em Atraso']
    hoje = date.today()
    rows = [
        [
            ln['cliente'], ln['titulo'], ln['valor'],
            ln['vencimento'].strftime('%d/%m/%Y') if ln['vencimento'] else '',
            ln['nosso_numero'] or '',
            (hoje - ln['vencimento']).days if ln['vencimento'] else '',
        ]
        for ln in linhas
    ]
    return _build_response(fmt, headers, rows, 'inadimplentes')
